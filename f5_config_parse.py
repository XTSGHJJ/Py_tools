# -*- encoding: utf-8 -*-

import re
import openpyxl
from os import getcwd, listdir
from openpyxl.styles import Alignment

def read_config_file(filepath):
    """读取配置文件内容"""
    try:
        with open(filepath, 'r') as file:
            return file.read()
    except IOError as e:
        print(f"Error reading file {filepath}: {e}")
        return None

def extract_data(pattern, text):
    """根据提供的正则表达式从文本中提取数据"""
    compiled_pattern = re.compile(pattern, re.MULTILINE)
    return re.findall(compiled_pattern, text)

def initialize_worksheet(wb, sheet_name):
    """初始化工作表，并设置标题和列头"""
    if sheet_name in wb.sheetnames:
        sheet_name += "_1"  # 防止重复的sheet名称
    ws = wb.create_sheet(title=sheet_name)
    headers = ['Pool Name', 'Pool Members', 
               'Virtual Server Name', 'Virtual Server IP', 'Virtual Server Port', 'Description',
               'SNAT Pool Name', 'SNAT Address','SNAT Translation Name','SNAT Translation Address',
               'VS Profile']
    for column, header in enumerate(headers, start=1):
        ws.cell(row=1, column=column, value=header)
    return ws

def remove_common_prefix(name):
    """移除名称中的公共前缀 '/Common/'"""
    return name.replace('/Common/', '')

def process_config_file(file_path, workbook):
    """处理单个配置文件"""
    config = read_config_file(file_path)
    if config is None:
        return
    
    sheet_name = file_path.split('/')[-1].split('.')[0]
    sheet = initialize_worksheet(workbook, sheet_name)

    # 定义用于数据提取的正则表达式
    pool_pattern = r'ltm pool (.+) {\n([\s\S]*?)\n}'
    vs_pattern = r'virtual (.+) {\n([\s\S]*?)\n}'
    snatpool_pattern = r'ltm snatpool (.+) {\n([\s\S]*?)\n}'
    snat_tran_pattern = r'ltm snat-translation (.+) {\n([\s\S]*?)\n}'
    snat_tran_addr = re.compile(r'address (\S+)')
    descr_pattern = re.compile(r'description (\S+)')
    profile_pattern = re.compile(r'profiles\s*{(\s*[\s\S]*?\s*)}\s*serverssl.*')
    profile_common = re.compile(r'/Common/(.*)\s*{')
    member_pattern = re.compile(r'(?:/Common/)([A-Fa-f0-9(:|.)]+[:.]\d+)', re.MULTILINE)
    destination_pattern = re.compile(r'destination ([^ ]+)[:.](\d+)')
    pool_ref_pattern = re.compile(r'pool (/\S+)')
    

    try:
        # 处理pool配置
        pools = extract_data(pool_pattern, config)
        row = 2
        for pool_name, members in pools:
            pool_name = remove_common_prefix(pool_name.strip())
            descr_match = descr_pattern.search(members)
            descr_str = descr_match.group(1) if descr_match else ''
            members = member_pattern.findall(members)
            member_str = "\n".join(m for m in members)
            sheet.cell(row=row, column=1, value=pool_name)
            sheet.cell(row=row, column=2, value=member_str)
            sheet.cell(row=row, column=6, value=descr_str).alignment = Alignment(wrap_text=True)
            row += 1

        # 处理vs配置
        virtual_servers = extract_data(vs_pattern, config)
        for vs_name, details in virtual_servers:
            vs_name = remove_common_prefix(vs_name.strip())
            dest_match = destination_pattern.search(details)
            pool_match = pool_ref_pattern.search(details)
            profile_mat = profile_pattern.findall(details)
            # print(details)
            # print(profile_mat)
            profile_match = profile_common.findall(profile_mat[0])
            # print(profile_match)
            profile_str = "\n".join(prof for prof in profile_match)
            # print(profile_str)
            if dest_match and pool_match:
                vs_ip = remove_common_prefix(dest_match.group(1))
                vs_port = dest_match.group(2)
                pool_name = remove_common_prefix(pool_match.group(1).strip())
                for i in range(2, sheet.max_row + 1):
                    if sheet.cell(row=i, column=1).value == pool_name:
                        sheet.cell(row=i, column=3, value=vs_name)
                        sheet.cell(row=i, column=4, value=vs_ip)
                        sheet.cell(row=i, column=5, value=vs_port)
                        sheet.cell(row=i, column=11, value=profile_str)
                        break

        # 处理snat pool配置            
        snatpools = extract_data(snatpool_pattern, config)
        row = 2
        for snat_name, members in snatpools:
            snat_name = remove_common_prefix(snat_name.strip())
            members = member_pattern.findall(members)
            member_str = "\n".join(m for m in members)
            sheet.cell(row=row, column=7, value=snat_name)
            sheet.cell(row=row, column=8, value=member_str).alignment = Alignment(wrap_text=True)
            row += 1
        # 处理snat translation配置   
        snat_tran = extract_data(snat_tran_pattern, config)
        row = 2
        for snat_tran_name, members in snat_tran:
            snat_name = remove_common_prefix(snat_tran_name.strip())
            members = snat_tran_addr.findall(members)
            member_str = "\n".join(m for m in members)
            sheet.cell(row=row, column=9, value=snat_name)
            sheet.cell(row=row, column=10, value=member_str).alignment = Alignment(wrap_text=True)
            row += 1

        # 调整列宽
        for col_letter, width in zip(['A', 'B', 'C', 'D', 'E', 'F'], [20, 40, 30, 20, 25, 30]):
            sheet.column_dimensions[col_letter].width = width

    except Exception as e:
        print(f"Error processing file {file_path}: {e}")

def main():
    """主函数, 执行配置文件的读取、数据提取和Excel文件的生成"""
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # 删除默认创建的sheet表

    pwd_dir = getcwd().replace('\\', '/')
    config_files = [f for f in listdir(pwd_dir) if f.endswith('.conf')] #匹配当前目录以.conf结尾的文件,可以修改
    if not config_files:
        print("No .conf files found in the current directory.")
        return

    for filename in config_files:
        file_path = pwd_dir + '/' + filename
        process_config_file(file_path, workbook)

    # 保存工作簿
    try:
        workbook.save('f5_ltm_config.xlsx')
        print("数据分析完成结果保存在f5_ltm_config.xlsx")
    except IOError as e:
        print(f"Error saving the workbook: {e}")

if __name__ == '__main__':
    main()
